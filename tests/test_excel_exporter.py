from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from app_indemnizaciones.domain.models import PeriodoLaborado
from app_indemnizaciones.services.cetil_models import (
    CetilEntidadEmpleadora,
    CetilExtractionResult,
    CetilMetadata,
    CetilTrabajador,
)
from app_indemnizaciones.services.excel_exporter import (
    build_ipc_export_info,
    build_liquidacion_filename,
    build_liquidacion_xlsx,
    safe_filename_document,
)
from app_indemnizaciones.services.ipc_loader import IpcRepository
from app_indemnizaciones.services.liquidacion_service import LiquidacionService


def test_exportador_incluye_columnas_kpis_datos_cetil_e_ipc():
    repo = IpcRepository.from_dataframe(
        pd.DataFrame(
            {
                "Fecha": ["1983/01/31", "1983/02/28", "2026/02/28"],
                "Índice de Precios al Consumidor (IPC)": [10, 20, 100],
            }
        )
    )
    resultado = LiquidacionService(repo).calcular(
        [PeriodoLaborado(date(1983, 9, 29), date(1983, 12, 31), Decimal("14000"), anio=1983)]
    )
    xlsx = build_liquidacion_xlsx(
        resultado,
        cetil_result=_cetil_result(),
        ipc_info=build_ipc_export_info(repo),
    )

    ws = load_workbook(BytesIO(xlsx), data_only=True)["Liquidación"]

    assert ws["A1"].value == "LIQUIDACIÓN DE INDEMNIZACIÓN SUSTITUTIVA DE VEJEZ"
    assert ws["A2"].value == "Nombre completo"
    assert ws["B2"].value == "Persona Prueba"
    assert ws["A5"].value == "Número CETIL"
    assert ws["B5"].value == "123456789"
    assert ws["D2"].value == "IPC actual/final"
    assert ws["E2"].value == 100
    assert ws["D4"].value == "Registros IPC"
    assert ws["E4"].value == 3
    assert "El año 1983 tiene 2 registros IPC" in ws["E5"].value

    headers = [ws.cell(row=13, column=column).value for column in range(1, 12)]
    assert headers == [
        "AÑO",
        "FECHA DESDE",
        "FECHA HASTA",
        "No. Días",
        "No. Sem.",
        "IBL Reportado",
        "% Apl.",
        "IPC Inicial",
        "IPC Actual",
        "Indexación IBC Mensual",
        "IBC Semanal Actualizado",
    ]
    assert ws["A17"].value == "DÍAS EN TOTAL"
    assert ws["A21"].value == "LIQUIDACIÓN DE APORTES"
    assert ws["A22"].value == "SBC x SC x PPC"


def test_exportador_escribe_formulas_en_celdas_calculadas():
    repo = IpcRepository.from_dataframe(
        pd.DataFrame(
            {
                "Fecha": ["1983/01/31", "1983/02/28", "2026/02/28"],
                "Índice de Precios al Consumidor (IPC)": [10, 20, 100],
            }
        )
    )
    resultado = LiquidacionService(repo).calcular(
        [PeriodoLaborado(date(1983, 9, 29), date(1983, 12, 31), Decimal("14000"), anio=1983)]
    )
    xlsx = build_liquidacion_xlsx(resultado, cetil_result=_cetil_result())

    ws = load_workbook(BytesIO(xlsx), data_only=False)["Liquidación"]

    assert ws["D14"].value == "=DAYS360(B14,C14)"
    assert ws["E14"].value == "=D14/7"
    assert ws["J14"].value == "=F14*(I14/H14)"
    assert ws["K14"].value == "=J14/4.345"
    assert ws["B17"].value == "=SUM(D14:D14)"
    assert ws["B18"].value == "=SUM(E14:E14)"
    assert ws["B19"].value == "=AVERAGE(G14:G14)"
    assert ws["B20"].value == "=AVERAGE(K14:K14)"
    assert ws["B21"].value == "=B20*B18*B19"


def test_nombre_archivo_incluye_documento():
    assert build_liquidacion_filename(_cetil_result()) == "Liquidacion_ISV_100.xlsx"
    assert safe_filename_document("88,152,324") == "88152324"


def test_nombre_archivo_sin_documento():
    assert build_liquidacion_filename() == "Liquidacion_ISV_sin_documento.xlsx"
    assert safe_filename_document(None) == "sin_documento"
    assert safe_filename_document("---") == "sin_documento"


def _cetil_result() -> CetilExtractionResult:
    return CetilExtractionResult(
        metadata=CetilMetadata(
            numero_cetil="123456789",
            ciudad_expedicion="BOGOTA",
            fecha_expedicion_cetil=date(2020, 1, 2),
        ),
        trabajador=CetilTrabajador(
            tipo_documento="CC",
            documento="100",
            nombre_completo="Persona Prueba",
            fecha_nacimiento=date(1960, 1, 1),
        ),
        entidad_empleadora=CetilEntidadEmpleadora(
            nombre_entidad_empleadora="Entidad Prueba",
            nit_entidad_empleadora="800100200",
        ),
    )
