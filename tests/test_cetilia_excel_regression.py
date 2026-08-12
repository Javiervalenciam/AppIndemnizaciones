from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from app_indemnizaciones.config import MONTHLY_TO_WEEKLY_DIVISOR
from app_indemnizaciones.domain.models import PeriodoLaborado, ResultadoLiquidacion
from app_indemnizaciones.services.excel_exporter import build_liquidacion_xlsx
from app_indemnizaciones.services.ipc_loader import IpcRepository
from app_indemnizaciones.services.liquidacion_service import LiquidacionService

CURRENT_BEHAVIOR_DUPLICATED_FORMULA = (
    "CURRENT_BEHAVIOR: Excel duplicates engine formulas; do not consolidate in this phase."
)
CENT = Decimal("0.01")


def _resultado() -> ResultadoLiquidacion:
    repo = IpcRepository.from_dataframe(
        pd.DataFrame(
            {
                "Fecha": [
                    "1983-01-31",
                    "1983-02-28",
                    "1984-01-31",
                    "1984-02-29",
                    "2026-04-30",
                ],
                "Índice de Precios al Consumidor (IPC)": [
                    Decimal("10"),
                    Decimal("20"),
                    Decimal("40"),
                    Decimal("60"),
                    Decimal("200"),
                ],
            }
        )
    )
    return LiquidacionService(repo).calcular(
        [
            PeriodoLaborado(
                fecha_inicio=date(1983, 9, 29),
                fecha_fin=date(1983, 12, 31),
                ibl_reportado=Decimal("14000.123"),
                anio=1983,
            ),
            PeriodoLaborado(
                fecha_inicio=date(1984, 1, 1),
                fecha_fin=date(1984, 11, 30),
                ibl_reportado=Decimal("11298.987"),
                anio=1984,
            ),
        ]
    )


def _xlsx_sheets(resultado: ResultadoLiquidacion):
    content = build_liquidacion_xlsx(resultado)
    formulas = load_workbook(BytesIO(content), data_only=False)["Liquidación"]
    values = load_workbook(BytesIO(content), data_only=True)["Liquidación"]
    return formulas, values


def _as_decimal(value: int | float) -> Decimal:
    return Decimal(str(value))


def test_resultado_liquidacion_matches_xlsx_cached_values() -> None:
    resultado = _resultado()
    _formulas, values = _xlsx_sheets(resultado)

    for excel_row, period in zip((14, 15), resultado.periodos, strict=True):
        assert values[f"B{excel_row}"].value.date() == period.fecha_inicio
        assert values[f"C{excel_row}"].value.date() == period.fecha_fin
        assert values[f"D{excel_row}"].value == period.dias
        assert values[f"E{excel_row}"].value == float(period.semanas)
        assert _as_decimal(values[f"F{excel_row}"].value) == period.ibl_reportado.quantize(CENT)
        assert values[f"G{excel_row}"].value == float(resultado.ppc)
        assert _as_decimal(values[f"H{excel_row}"].value) == period.ipc_inicial
        assert _as_decimal(values[f"I{excel_row}"].value) == period.ipc_actual
        assert _as_decimal(values[f"J{excel_row}"].value) == period.ibc_actualizado.quantize(CENT)
        assert _as_decimal(values[f"K{excel_row}"].value) == (
            period.ibc_semanal_actualizado.quantize(CENT)
        )

    assert values["B18"].value == resultado.total_dias
    assert values["B19"].value == float(resultado.sc)
    assert values["B20"].value == float(resultado.ppc)
    assert _as_decimal(values["B21"].value) == resultado.sbc.quantize(CENT)
    assert _as_decimal(values["B22"].value) == resultado.isv.quantize(CENT)


def test_current_excel_formulas_duplicate_engine_formulas_by_design() -> None:
    resultado = _resultado()
    formulas, _values = _xlsx_sheets(resultado)

    assert formulas["D14"].value == "=DAYS360(B14,C14)", CURRENT_BEHAVIOR_DUPLICATED_FORMULA
    assert formulas["E14"].value == "=D14/7", CURRENT_BEHAVIOR_DUPLICATED_FORMULA
    assert formulas["J14"].value == "=F14*(I14/H14)", CURRENT_BEHAVIOR_DUPLICATED_FORMULA
    assert formulas["K14"].value == (
        f"=J14/{MONTHLY_TO_WEEKLY_DIVISOR}"
    ), CURRENT_BEHAVIOR_DUPLICATED_FORMULA
    assert formulas["B18"].value == "=SUM(D14:D15)", CURRENT_BEHAVIOR_DUPLICATED_FORMULA
    assert formulas["B19"].value == "=SUM(E14:E15)", CURRENT_BEHAVIOR_DUPLICATED_FORMULA
    assert formulas["B20"].value == "=AVERAGE(G14:G15)", CURRENT_BEHAVIOR_DUPLICATED_FORMULA
    assert formulas["B21"].value == "=AVERAGE(K14:K15)", CURRENT_BEHAVIOR_DUPLICATED_FORMULA
    assert formulas["B22"].value == "=B21*B19*B20", CURRENT_BEHAVIOR_DUPLICATED_FORMULA
